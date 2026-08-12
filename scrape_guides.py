import requests
import re
import time
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ใช้ GitHub API เดินหาไฟล์ .md ทั้งหมดในโฟลเดอร์ content/guides อัตโนมัติ
API_BASE = "https://api.github.com/repos/docker/docs/contents/"
RAW_BASE = "https://raw.githubusercontent.com/docker/docs/main/"
START_PATH = "content/guides"


def list_md_files(path):
    """เดินเข้าทุกโฟลเดอร์ย่อย เก็บ path ของไฟล์ .md ทั้งหมด"""
    md_files = []
    try:
        r = requests.get(API_BASE + path, timeout=15)
        if r.status_code != 200:
            print(f"  เข้าไม่ได้: {path} (status {r.status_code})")
            return md_files
        items = r.json()
        for item in items:
            if item["type"] == "file" and item["name"].endswith(".md"):
                md_files.append(item["path"])
            elif item["type"] == "dir":
                time.sleep(0.3)  # กัน rate limit
                md_files.extend(list_md_files(item["path"]))  # เข้าโฟลเดอร์ย่อย
    except Exception as e:
        print(f"  error {path}: {e}")
    return md_files


def clean_markdown(text):
    text = re.sub(r"^---.*?---", "", text, count=1, flags=re.DOTALL)
    text = re.sub(r"<!--.*?-->", "", text, flags=re.DOTALL)
    text = re.sub(r"\{\{.*?\}\}", "", text, flags=re.DOTALL)  # ตัด template tag
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"```[a-zA-Z]*\n", "", text)
    text = text.replace("```", "")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def split_by_heading(text):
    chunks = []
    parts = re.split(r"\n(#{2,3})\s+", text)
    if parts[0].strip():
        first_title = parts[0].split("\n")[0].replace("#", "").strip()[:60]
        chunks.append((first_title or "บทนำ", parts[0].strip()))
    i = 1
    while i < len(parts) - 1:
        heading = parts[i + 1].split("\n")[0].strip()
        body = "\n".join(parts[i + 1].split("\n")[1:]).strip()
        if body and len(body) > 40:
            chunks.append((heading[:60], body))
        i += 2
    return chunks


# ===== เดินหาไฟล์ทั้งหมด =====
print("กำลังหาไฟล์ .md ทั้งหมดใน content/guides ...")
all_files = list_md_files(START_PATH)
print(f"เจอไฟล์ทั้งหมด {len(all_files)} ไฟล์\n")

# ===== ดึงเนื้อหาแต่ละไฟล์ + กันซ้ำ =====
rows = []
seen = set()   # กันเนื้อหาซ้ำ
idx = 1
dup = 0

for i, page in enumerate(all_files, 1):
    url = RAW_BASE + page
    try:
        r = requests.get(url, timeout=15)
        if r.status_code != 200:
            continue
        text = clean_markdown(r.text)
        source = page.replace("content/guides/", "").replace(".md", "")
        for title, body in split_by_heading(text):
            body = body[:1500]
            key = body[:100]  # ใช้ 100 ตัวแรกเช็คซ้ำ
            if key in seen:
                dup += 1
                continue
            seen.add(key)
            rows.append([idx, source, title, body])
            idx += 1
        if i % 10 == 0:
            print(f"  ดึงแล้ว {i}/{len(all_files)} ไฟล์ ... ได้ {len(rows)} chunks")
        time.sleep(0.3)
    except Exception as e:
        print(f"  error {page}: {e}")

print(f"\nรวม {len(rows)} chunks (ตัดซ้ำออก {dup} อัน)")

# ===== เขียน Excel ลง knowledge/ =====
os.makedirs("knowledge", exist_ok=True)
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "knowledge"
ws.append(["id", "แหล่งที่มา", "หัวข้อ", "เนื้อหา"])
for row in rows:
    ws.append(row)

fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
hfont = Font(name="Arial", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, 5):
    cell = ws.cell(1, c)
    cell.fill = fill; cell.font = hfont
    cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
for r in range(2, ws.max_row + 1):
    for c in range(1, 5):
        cell = ws.cell(r, c)
        cell.font = Font(name="Arial", size=10); cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 90

wb.save("knowledge/docker_guides_dataset.xlsx")
print("บันทึกลง knowledge/docker_guides_dataset.xlsx เสร็จ")
print("แหล่งอ้างอิง: Docker Documentation - Guides (github.com/docker/docs)")